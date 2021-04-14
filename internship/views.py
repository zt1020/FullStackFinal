"""
views.py
Contributors: Zachary, Purnya
Updated : 04/03/2021
"""

from openpyxl import load_workbook
from django.contrib.auth.decorators import login_required
from django.views.generic import TemplateView
from django.contrib import messages
from faker import Faker
from django.shortcuts import redirect  # pylint: disable=C0412
from django.contrib.auth.forms import AuthenticationForm # pylint: disable=C0412
from django.contrib.auth import login,logout,authenticate # pylint: disable=C0412
from django.contrib.auth.models import User, Group # pylint: disable=C0412
from django.shortcuts import (get_object_or_404,render,HttpResponseRedirect)
from internship.models import Student,Internship,InternshipAssignment
from .models import Student, InternshipAssignment,Internship
from .forms import StudentSearchForm,UpdateInternshipAssignmentForm
from .forms import InternshipSearchForm,CreateUserForm
from .forms import InternshipAssignmentSearchForm,StudentForm,InternshipForm
from .roles import allowed_users
from django.contrib import messages


f_data = Faker()



def login_request(request):
    """
    method for login page
    """
    if request.method == 'POST':
        form = AuthenticationForm(request=request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None: # pylint: disable=R1705
                login(request, user)
                # messages.info(request, f"You are now logged in as {username}")
                return redirect('home')
        #     else:
        #         messages.error(request, "Invalid username or password.")
        #else:
            #messages.error(request, "Invalid username or password.")
    #form = AuthenticationForm()
    if form.is_valid() == False:
        messages.error(request, "Invalid username or password.")
        return render(request = request,template_name = "accounts/login.html",context={"form":form})


def logout_request(request):
    """
    log out page
    """
    logout(request)
    messages.info(request,"Logged out successfully!")
    return render(request,"accounts/logout.html")


def register_page(request):
    """
    register page
    """
    form = CreateUserForm()
    if request.method == "POST":
        form = CreateUserForm(request.POST)
        if form.is_valid():
            user = form.save()
            group = form.cleaned_data['group']
            group = Group.objects.get(name=group)
            user.groups.add(group)
    context = {'form' : form}
    return render(request, 'accounts/register.html', context)

@allowed_users(allowed_roles=['Instructor'])
def create_student(request):
    """
    create students
    """
    form = StudentForm()
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/')
    return render(request, 'accounts/insert_view.html', {'form':form})

@allowed_users(allowed_roles=['Instructor'])
def update_student(request,pk):
    """
    update student
    """
    context ={}

    obj = get_object_or_404(Student, pk = pk)


    student_form = StudentForm(request.POST or None,instance=obj)

    if student_form.is_valid():

        student_form.save()

        messages.success(request, 'Details updated successfully! \
            Go to Student page to view the updates.')
    context["student_form"] = student_form

    return render(request, "accounts/update_view.html", context)

@allowed_users(allowed_roles=['Instructor'])
def create_internship(request):
    """
    create internship
    """
    form = InternshipForm()
    if request.method == 'POST':
        form = InternshipForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/')
    return render(request, 'accounts/insert_view_internship.html',
                  {'form':form})


@allowed_users(allowed_roles=['Instructor'])
def create_internshipassignment(request):
    """
    create internship Assignment
    """
    form = UpdateInternshipAssignmentForm()
    if request.method == 'POST':
        form = UpdateInternshipAssignmentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/')
    return render(request, 'accounts/insert_view_internship.html',
                  {'form':form})


@allowed_users(allowed_roles=['Instructor'])
def update_internship(request,pk):
    """
    update internship
    """
    context ={}

    obj = get_object_or_404(Internship, pk = pk)


    internship_form = InternshipForm(request.POST or None,instance=obj)

    if internship_form.is_valid():

        internship_form.save()

        messages.success(request, 'Details updated successfully! \
            Go to Internship page to view the updates.')
    context["internship_form"] = internship_form

    return render(request, "accounts/update_view_internship.html", context)

@allowed_users(allowed_roles=['Instructor'])
def update_internshipassignment(request,pk):
    """
    update internship Assignment
    """
    context ={}

    obj = get_object_or_404(InternshipAssignment, pk = pk)

    internshipAssignment_form = UpdateInternshipAssignmentForm(request.POST or None,instance=obj)

    if internshipAssignment_form.is_valid():

        internshipAssignment_form.save()

        messages.success(request, 'Details updated successfully! \
            Go to Internship Assignment page to view the updates.')

    context["internshipAssignment_form"] = internshipAssignment_form

    return render(request, "accounts/update_view_internshipassignment.html",
                  context)

@allowed_users(allowed_roles=['Instructor'])
def delete_student(request, pk):
    """
    delete student
    """
    context ={}
    obj = get_object_or_404(Student, pk = pk)
    if request.method =="POST":
        obj.delete()
        return HttpResponseRedirect("/")
    return render(request, "accounts/delete_view.html", context)

@allowed_users(allowed_roles=['Instructor'])
def delete_internship(request, pk):
    """
    delete internship
    """
    context ={}
    obj = get_object_or_404(Internship, pk = pk)
    if request.method =="POST":
        obj.delete()
        return HttpResponseRedirect("/")
    return render(request, "accounts/delete_view_internship.html", context)

@allowed_users(allowed_roles=['Instructor'])
def delete_internshipassignment(request, pk):
    """
    delete internship Assignment
    """
    context ={}
    obj = get_object_or_404(InternshipAssignment, pk = pk)
    if request.method =="POST":
        obj.delete()
        return HttpResponseRedirect("/")
    return render(request, "accounts/delete_view_internshipassignment.html",
                  context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['Instructor'])
def import_file(request):
    """
    import the file
    """
    if request.method=='POST':
        try:
            file = request.FILES['document']
            messages.success(request, 'Form submission successful')
            import_data(file)
        except KeyError:
            messages.info(request, 'No file selected')
    return render(request, 'import.html')


def import_data(request):
    """
    loads the workbook and write functions to import the data
    """
    work_book = load_workbook(filename = request, data_only=True)
    sheet = work_book['sample-internship-data']
    import_student(sheet)
    import_internship(sheet)
    import_internshipassignment(sheet)

def import_student(sheet):
    """
    importing student and generating fake data
    """
    for i in range(2, sheet.max_row+1):
        student_id=str(i-1)
        unh_id=str(i-1)
        last_name=f_data.last_name()
        first_name=f_data.first_name()
        major=sheet['D'+str(i)].value
        school_email=f_data.email()
        degree=sheet['E'+str(i)].value
        linkedin=f_data.url()

        import_s = Student(
            student_id=student_id,
            unh_id=unh_id,last_name=last_name,first_name=first_name,
            school_email=  school_email, major = major,
            degree=  degree, linkedin = linkedin
        )
        import_s.save()

def import_internship(sheet):
    """
    importing internship and generating fake data
    """
    for i in range(2, sheet.max_row+1):

        internship_id=str(i-1)
        position=sheet['N'+str(i)].value
        pay=sheet['O'+str(i)].value
        organization_name=sheet['R'+str(i)].value
        organization_url=f_data.url()
        organization_address=sheet['T'+str(i)].value
        supervisor_name=sheet['X'+str(i)].value
        supervisor_position=sheet['Y'+str(i)].value
        supervisor_email=sheet['Z'+str(i)].value
        supervisor_phone=sheet['AA'+str(i)].value


        import_i = Internship(
            internship_id=internship_id,
            position=position,
            pay=pay,
            organization_name=organization_name,
            organization_url=organization_url,
            organization_address=organization_address,
            supervisor_name=supervisor_name,
            supervisor_position=supervisor_position,
            supervisor_email=supervisor_email,
            supervisor_phone=supervisor_phone
        )
        import_i.save()

def import_internshipassignment(sheet):
    """
    importing InternshipAssignment and generating fake data
    """
    for i in range(2, sheet.max_row+1):

        course_id=sheet['I'+str(i)].value
        student_credits=sheet['J'+str(i)].value
        semester=sheet['K'+str(i)].value
        year=sheet['L'+str(i)].value
        instructor=sheet['M'+str(i)].value
        start_date=f_data.date()
        end_date=f_data.date()

        import_ia = InternshipAssignment(
            course_id=course_id,student_credits=student_credits,
            semester=semester,
            year=year,
            instructor=instructor,
            start_date=start_date,end_date=end_date
        )
        import_ia.save()


class HomepageView(TemplateView):
    """
    This creates class based HomepageView
    """
    template_name = 'home.html'

@login_required(login_url='login')
@allowed_users(allowed_roles=['Instructor'])
def display_students(request):
    """
    display dropdown fields for first_name and last_name in Student
    """
    button = "students"
    student_items = Student.objects.all() # pylint: disable=E1101
    form = StudentSearchForm(request.POST or None)
    first_name=Student.objects.all() # pylint: disable=E1101
    last_name=Student.objects.all() # pylint: disable=E1101
    context = {
        'button' : button,
        'student_items' : student_items,
        'form' : form,
        'first_name':first_name,
        'last_name':last_name
    }
    if request.method == 'POST':
        student_items = Student.objects.filter( # pylint: disable=E1101
            first_name__icontains=form['first_name'].value(),
            last_name__icontains=form['last_name'].value()
            )
        context = {
            "student_items" : student_items,
            "form": form
        }
    return render(request, 'display_students.html', context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['Instructor','Current','Incoming'])
def display_internship(request):
    """
    display dropdown fields for organization_name and supervisor_name
    in Internship
    """
    button = "Internship"
    internship_items = Internship.objects.all() # pylint: disable=E1101
    form = InternshipSearchForm(request.POST or None)
    organization_name=Internship.objects.all() # pylint: disable=E1101
    supervisor_name=Internship.objects.all() # pylint: disable=E1101
    context = {
        'button' : button,
        'internship_items' : internship_items,
        'form' : form,
        'organization_name' : organization_name,
        'supervisor_name' : supervisor_name
    }

    if request.method == 'POST':
        internship_items = Internship.objects.filter( # pylint: disable=E1101
            organization_name__icontains=form['organization_name'].value(),
            supervisor_name__icontains=form['supervisor_name'].value()
            )
        context = {
            "internship_items" : internship_items,
            "form": form
        }
    return render(request, 'display_internships.html', context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['Instructor'])
def display_internshipassignment(request):
    """
    display dropdown fields for semester and year
    in Internship Assignment
    """
    button = "Internship Assignment"
    internshipassignment_items = InternshipAssignment.objects.all() # pylint: disable=E1101
    form = InternshipAssignmentSearchForm(request.POST or None)
    semester = InternshipAssignment.objects.all() # pylint: disable=E1101
    year = InternshipAssignment.objects.all() # pylint: disable=E1101
    context = {
        'button' : button,
        'internshipassignment_items' : internshipassignment_items,
        'form' : form,
        'semester' : semester,
        'year' : year
    }
    if request.method == 'POST':
        internshipassignment_items = InternshipAssignment.objects.filter( # pylint: disable=E1101
            semester__icontains=form['semester'].value(),
            year__icontains=form['year'].value()
            )
        context = {
            "internshipassignment_items" : internshipassignment_items,
            "form": form
        }
    return render(request, 'display_internshipassignment.html', context)

@allowed_users(allowed_roles=['Instructor'])
def remove_data(request):
    """
    Removing the data from all database tables
    """
    Student.objects.all().delete() #pylint: disable = no-member
    InternshipAssignment.objects.all().delete() #pylint: disable = no-member
    Internship.objects.all().delete() #pylint: disable = no-member
    return render(request, 'home.html')
